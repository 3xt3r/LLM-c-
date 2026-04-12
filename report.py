from __future__ import annotations

import datetime
import json
from pathlib import Path

from models import ComponentRecord

_CLASSIFICATIONS = [
    "vendored / in-tree",
    "external system dependency",
    "unresolved",
    "not a library",
]


# ─────────────────────────────────────────────────────────────────────────────
# Markdown
# ─────────────────────────────────────────────────────────────────────────────

def _paths(items: list, limit: int = 3) -> str:
    vals = []
    for item in items[:limit]:
        vals.append(f"{item.file}:{item.line}" if item.line else item.file)
    return " · ".join(vals)


def build_markdown(records: list[ComponentRecord]) -> str:
    summary = {cls: 0 for cls in _CLASSIFICATIONS}
    for rec in records:
        summary[rec.classification] = summary.get(rec.classification, 0) + 1

    lines: list[str] = []
    lines.append("# Dependency analysis report\n")
    lines.append("## Summary\n")
    for cls in _CLASSIFICATIONS:
        lines.append(f"- **{cls}**: {summary.get(cls, 0)}")
    lines.append("")

    interesting = [r for r in records if r.classification != "not a library"]
    lines.append("## Component table\n")
    lines.append(
        "| Component | Classification | Evidence level | Confidence | Optional/platform | Why | Link evidence | Source path |"
    )
    lines.append("|---|---|---|---|---|---|---|---|")
    for rec in interesting:
        link = _paths(rec.final_link_evidence)
        src  = _paths(rec.in_tree_source_evidence)
        why  = rec.why.replace("|", "/")
        lines.append(
            f"| {rec.normalized_name} | {rec.classification} | {rec.evidence_level} | "
            f"{rec.confidence} | {rec.optional_or_platform_specific} | {why} | {link} | {src} |"
        )
    lines.append("")

    def section(title: str, cls: str) -> None:
        subset = [r for r in records if r.classification == cls]
        lines.append(f"## {title}\n")
        if not subset:
            lines.append("_None_\n")
            return
        for rec in subset:
            missing = ""
            if rec.missing_evidence:
                missing = " — missing: " + "; ".join(rec.missing_evidence)
            lines.append(
                f"- **{rec.normalized_name}** — {rec.evidence_level}; "
                f"{rec.confidence} confidence. {rec.why}{missing}"
            )
        lines.append("")

    section("Vendored / in-tree", "vendored / in-tree")
    section("External system dependencies", "external system dependency")
    section("Unresolved (needs manual review)", "unresolved")

    system = [r for r in records if r.classification == "not a library"]
    lines.append(f"## Not libraries (system/compiler headers) — {len(system)} total\n")
    if system:
        lines.append("<details><summary>Expand list</summary>\n")
        for rec in system:
            lines.append(f"- {rec.normalized_name}")
        lines.append("\n</details>\n")

    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers for Russian labels
# ─────────────────────────────────────────────────────────────────────────────

def _action_for(rec: ComponentRecord) -> str:
    """Return a recommended next action in Russian based on classification + confidence."""
    cls  = rec.classification
    conf = rec.confidence

    if cls == "not a library":
        return "—"

    if cls == "vendored / in-tree":
        if conf == "high":
            return "Нет действий — компонент встроен в репозиторий."
        return "Уточнить источник — исходники найдены, но участие в сборке не подтверждено. Проверить Makefile.am / CMakeLists.txt."

    if cls == "external system dependency":
        if conf == "high":
            opt = rec.optional_or_platform_specific
            if opt.startswith("yes:"):
                return f"Опционально — устанавливать только если нужен этот бэкенд ({opt.replace('yes: ', '')})."
            return "Установить на хост перед сборкой (apt install / brew install / vcpkg и т.д.)."
        return "Проверить вручную — уверенность средняя. Открыть лист Улики и проверить файлы сборки."

    if cls == "unresolved":
        if rec.missing_evidence:
            return "Проверить вручную — " + rec.missing_evidence[0]
        return "Проверить вручную — недостаточно улик для классификации."

    return "Проверить вручную."


def _ev_kind_ru(kind: str) -> str:
    """Translate evidence kind to Russian."""
    return {
        "include":                 "include (заголовок)",
        "vendored_dir":            "исходники в репо (директория)",
        "vendored_source":         "исходники в репо",
        "build_detection":         "обнаружение в сборке",
        "build_integration":       "интеграция в сборку",
        "build_integration_cmake": "интеграция в сборку (CMake)",
        "final_link":              "финальная линковка",
        "final_link_cmake":        "финальная линковка (CMake)",
        "platform_note":           "платформенная заметка",
    }.get(kind, kind.replace("_", " "))


# ─────────────────────────────────────────────────────────────────────────────
# Excel comments — подсказки к заголовкам колонок
# ─────────────────────────────────────────────────────────────────────────────

# Комментарии к каждой колонке инвентаря и сегментных листов
_COL_COMMENTS: dict[str, str] = {
    "Версия": (
        "Версия компонента, извлечённая из репозитория.\n"
        "Источники (в порядке приоритета):\n"
        "1. Файлы в vendored-директории (VERSION.txt, CMakeLists.txt и т.д.)\n"
        "2. Сниппеты build-улик (find_package, AC_INIT, PKG_CHECK_MODULES)\n"
        "3. Имена пакетов (yaml-0.1, pcre2-8 >= 10.30)\n"
        "— означает что версия не найдена."
    ),
    "Уверенность в версии": (
        "Насколько надёжно найдена версия:\n"
        "• HIGH — из надёжного источника (AC_INIT, find_package, VERSION файл)\n"
        "• MEDIUM — из заголовка или переменной сборки\n"
        "• LOW — из имени пакета или косвенного упоминания"
    ),
    "CPE": (
        "Common Platform Enumeration — стандартный идентификатор для CVE-сканеров.\n"
        "Формат: cpe:2.3:a:vendor:product:version:*:*:*:*:*:*:*\n"
        "* в поле версии означает что версия не определена.\n"
        "Используется для поиска уязвимостей в NVD, OSV и других базах."
    ),
    "Компонент": (
        "Нормализованное имя библиотеки.\n"
        "Пример: libhtp, libpcap, zlib.\n"
        "Всегда в нижнем регистре без версии."
    ),
    "Классификация": (
        "Итоговая категория компонента:\n"
        "• vendored / in-tree — исходники внутри репозитория\n"
        "• external system dependency — нужно установить на хост\n"
        "• unresolved — не удалось определить\n"
        "• not a library — системный заголовок"
    ),
    "Уровень улик": (
        "Качество доказательств классификации:\n"
        "• in-tree source + build participation — сильнейший\n"
        "• confirmed linked — явная линковка (-lNAME)\n"
        "• build-integrated — в переменных сборки\n"
        "• probe only — только проверка наличия\n"
        "• insufficient evidence — данных нет"
    ),
    "Уверенность": (
        "Уровень уверенности агента:\n"
        "• HIGH — несколько сильных улик, ручная проверка не нужна\n"
        "• MEDIUM — одна из улик слабая или отсутствует\n"
        "• LOW — только слабые улики, обязательно проверить вручную"
    ),
    "Опционально / платформа": (
        "Является ли зависимость опциональной или платформозависимой.\n"
        "— означает: обязательна на всех платформах.\n"
        "Примеры: 'опциональный бэкенд DPDK', 'специфично для Windows'"
    ),
    "Обоснование": (
        "Объяснение почему агент принял такое решение.\n"
        "Читай когда не согласен с классификацией.\n"
        "Основано на найденных уликах."
    ),
    "Чего не хватает": (
        "Какие улики отсутствуют для более уверенного ответа.\n"
        "Если заполнено — агент не уверен.\n"
        "Пример: 'Требуется LDADD / target_link_libraries'"
    ),
    "Действие": (
        "Рекомендация что делать с этим компонентом:\n"
        "• Нет действий — всё ок, встроен в репо\n"
        "• Установить на хост — внешняя зависимость\n"
        "• Проверить вручную — низкая уверенность\n"
        "• Уточнить источник — исходники есть, сборка не подтверждена"
    ),
    "Источник обнаружения": (
        "Как агент впервые нашёл этот компонент:\n"
        "• in-tree source — найдена директория с .c/.cpp\n"
        "• build config — найден в configure.ac / CMakeLists.txt\n"
        "• header only — только через #include\n"
        "• in-tree source + build config — и то и другое"
    ),
    # Evidence sheet
    "Тип улики": (
        "Откуда взята эта строка-доказательство:\n"
        "• include (заголовок) — строка #include в .c/.cpp файле\n"
        "• финальная линковка — target_link_libraries / -lNAME\n"
        "• интеграция в сборку — AC_SUBST, LDADD, AM_CONDITIONAL\n"
        "• обнаружение в сборке — AC_CHECK_LIB, PKG_CHECK_MODULES\n"
        "• исходники в репо — директория с .c/.cpp файлами\n"
        "• платформенная заметка — упоминание платформы рядом с именем"
    ),
    "Файл": (
        "Полный путь к файлу где найдена эта улика.\n"
        "Можно открыть в редакторе чтобы проверить контекст."
    ),
    "Строка": (
        "Номер строки в файле где найдена улика.\n"
        "— означает что улика на уровне директории (не конкретной строки)."
    ),
    "Фрагмент кода": (
        "Сама строка кода или конфига где найдена улика.\n"
        "Первые 300 символов строки."
    ),
    # Optional backends
    "Тип бэкенда": (
        "Описание для чего нужен этот опциональный компонент.\n"
        "Пример: 'опциональный бэкенд DPDK' — высокопроизводительный захват пакетов.\n"
        "Если компонент не нужен — его можно не устанавливать."
    ),
    # System headers
    "Заголовок / интринсик": (
        "Имя системного заголовка или интринсика компилятора.\n"
        "Это НЕ сторонние зависимости — они входят в компилятор или ОС."
    ),
    "Категория": (
        "Тип системного заголовка:\n"
        "• стандартная библиотека — stdio.h, string.h, vector и т.д.\n"
        "• интринсик компилятора — SIMD заголовки (emmintrin.h и т.д.)\n"
        "• заголовок ядра / платформы — linux/bpf.h, sys/socket.h\n"
        "• Windows SDK — windows.h, winsock2.h"
    ),
    "Причина": (
        "Почему этот заголовок помечен как 'не библиотека'.\n"
        "Такие заголовки присутствуют в любой среде разработки\n"
        "и не требуют отдельной установки."
    ),
    # Dashboard
    "Классификация": (
        "Итоговая категория компонента:\n"
        "• vendored / in-tree — исходники внутри репозитория\n"
        "• external system dependency — нужно установить на хост\n"
        "• unresolved — не удалось определить\n"
        "• not a library — системный заголовок"
    ),
    "Высокая": "Количество компонентов с высокой уверенностью (HIGH).\nМожно доверять без ручной проверки.",
    "Средняя": "Количество компонентов со средней уверенностью (MEDIUM).\nРекомендуется беглая проверка.",
    "Низкая":  "Количество компонентов с низкой уверенностью (LOW).\nОбязательно проверить вручную.",
    "Итого":   "Суммарное количество компонентов в этой категории.",
}


def _add_comment(cell, text: str, author: str = "dep-agent") -> None:
    """Attach a Russian tooltip comment to a cell."""
    from openpyxl.comments import Comment
    comment = Comment(text, author)
    comment.width  = 320
    comment.height = max(60, min(200, text.count("\n") * 18 + 40))
    cell.comment = comment


# ─────────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────────

def build_xlsx(records: list[ComponentRecord], out_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.chart import PieChart, Reference
        from openpyxl.chart.series import DataPoint
    except ImportError:
        print("  [xlsx] openpyxl not installed — skipping. Run: pip install openpyxl")
        return

    C_DARK        = "1A2B3C"
    C_BLUE_HEAD   = "1A5276"
    C_BLUE_LIGHT  = "D6EAF8"
    C_GREEN_HEAD  = "1E8449"
    C_GREEN_LIGHT = "D5F5E3"
    C_AMBER_HEAD  = "9C6400"
    C_AMBER_LIGHT = "FDEBD0"
    C_RED_HEAD    = "922B21"
    C_RED_LIGHT   = "FADBD8"
    C_GRAY_HEAD   = "2C3E50"
    C_GRAY_LIGHT  = "F2F3F4"
    C_WHITE       = "FFFFFF"
    C_BORDER      = "BFC9CA"

    CLS_BG = {
        "vendored / in-tree":         C_GREEN_LIGHT,
        "external system dependency":  C_BLUE_LIGHT,
        "unresolved":                  C_AMBER_LIGHT,
        "not a library":               C_GRAY_LIGHT,
    }
    CLS_FG = {
        "vendored / in-tree":         C_GREEN_HEAD,
        "external system dependency":  C_BLUE_HEAD,
        "unresolved":                  C_AMBER_HEAD,
        "not a library":               C_GRAY_HEAD,
    }
    CONF_BG = {"high": C_GREEN_LIGHT, "medium": C_AMBER_LIGHT, "low": C_RED_LIGHT}
    CONF_FG = {"high": C_GREEN_HEAD,  "medium": C_AMBER_HEAD,  "low": C_RED_HEAD}
    EV_BG = {
        "vendored_dir":            C_GREEN_LIGHT,
        "vendored_source":         C_GREEN_LIGHT,
        "include":                 C_GRAY_LIGHT,
        "build_detection":         C_AMBER_LIGHT,
        "build_integration":       C_BLUE_LIGHT,
        "build_integration_cmake": C_BLUE_LIGHT,
        "final_link":              "EAF2FF",
        "final_link_cmake":        "EAF2FF",
        "platform_note":           "F5EEF8",
    }
    EV_FG = {
        "vendored_dir":            C_GREEN_HEAD,
        "vendored_source":         C_GREEN_HEAD,
        "include":                 "5D6D7E",
        "build_detection":         C_AMBER_HEAD,
        "build_integration":       C_BLUE_HEAD,
        "build_integration_cmake": C_BLUE_HEAD,
        "final_link":              "154360",
        "final_link_cmake":        "154360",
        "platform_note":           "6C3483",
    }

    def _side(): return Side(border_style="thin", color=C_BORDER)
    def _border(): return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
    def _fill(h): return PatternFill("solid", fgColor=h)
    def _font(bold=False, size=10, color="000000", italic=False):
        return Font(bold=bold, size=size, color=color, name="Arial", italic=italic)
    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def sc(cell, bg=C_WHITE, fg="000000", bold=False, size=10,
           h="left", v="center", wrap=False, italic=False, border=True):
        cell.fill      = _fill(bg)
        cell.font      = _font(bold=bold, size=size, color=fg, italic=italic)
        cell.alignment = _align(h=h, v=v, wrap=wrap)
        if border:
            cell.border = _border()

    def col(n):
        from openpyxl.utils import get_column_letter
        return get_column_letter(n)

    wb = Workbook()
    interesting  = [r for r in records if r.classification != "not a library"]
    optional_rec = [r for r in records if r.optional_or_platform_specific.startswith("yes")]
    system_recs  = [r for r in records if r.classification == "not a library"]
    counts = {cls: sum(1 for r in records if r.classification == cls) for cls in _CLASSIFICATIONS}

    # ── Sheet 1: Dashboard ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    ws.merge_cells("B2:M2")
    c = ws["B2"]
    c.value = "Анализ зависимостей C/C++"
    sc(c, bg=C_DARK, fg=C_WHITE, bold=True, size=16, h="left", border=False)
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:M3")
    c = ws["B3"]
    c.value = f"Сформировано dep-agent  ·  {datetime.date.today().strftime('%d.%m.%Y')}"
    sc(c, bg=C_DARK, fg="7F8C8D", size=9, h="left", border=False)
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 10

    cards = [
        ("Встроены в репозиторий", counts["vendored / in-tree"],        C_GREEN_HEAD, C_GREEN_LIGHT),
        ("Внешние зависимости",    counts["external system dependency"], C_BLUE_HEAD,  C_BLUE_LIGHT),
        ("Не определено",          counts["unresolved"],                C_AMBER_HEAD, C_AMBER_LIGHT),
        ("Всего компонентов",      len(records),                        C_GRAY_HEAD,  C_GRAY_LIGHT),
    ]
    card_cols = [2, 5, 8, 11]
    for i, (label, value, hc, lc) in enumerate(cards):
        cs = card_cols[i]; ce = cs + 2
        for r in range(5, 11):
            for cc in range(cs, ce + 1):
                ws.cell(row=r, column=cc).fill   = _fill(lc)
                ws.cell(row=r, column=cc).border = Border()
        ws.merge_cells(f"{col(cs)}5:{col(ce)}7")
        ws.merge_cells(f"{col(cs)}8:{col(ce)}9")
        c = ws[f"{col(cs)}5"]; c.value = value
        sc(c, bg=lc, fg=hc, bold=True, size=32, h="center", v="center", border=False)
        c = ws[f"{col(cs)}8"]; c.value = label
        sc(c, bg=lc, fg=hc, bold=True, size=10, h="center", v="center", border=False)

    for r in [5,6,7,8,9,10]:
        ws.row_dimensions[r].height = 14 if r != 8 else 20
    ws.row_dimensions[11].height = 10

    ws["B12"].value = "Разбивка по уровню уверенности"
    sc(ws["B12"], fg=C_DARK, bold=True, size=11, border=False)
    ws.row_dimensions[12].height = 22

    hdr = ["Классификация", "Высокая", "Средняя", "Низкая", "Итого"]
    cw  = [28, 10, 10, 10, 10]
    for j, (h, w) in enumerate(zip(hdr, cw)):
        ws.column_dimensions[col(j+2)].width = w
        c = ws.cell(row=13, column=j+2, value=h)
        sc(c, bg=C_GRAY_HEAD, fg=C_WHITE, bold=True, size=10, h="center")
        if h in _COL_COMMENTS:
            _add_comment(c, _COL_COMMENTS[h])
    ws.row_dimensions[13].height = 18

    for i, cls in enumerate(["vendored / in-tree", "external system dependency", "unresolved"]):
        row = 14 + i
        subset = [r for r in records if r.classification == cls]
        for j, v in enumerate([cls,
                                sum(1 for r in subset if r.confidence=="high"),
                                sum(1 for r in subset if r.confidence=="medium"),
                                sum(1 for r in subset if r.confidence=="low"),
                                len(subset)]):
            c = ws.cell(row=row, column=j+2, value=v)
            sc(c, bg=CLS_BG.get(cls, C_WHITE), fg=CLS_FG.get(cls, "000000"),
               bold=(j==0), size=10, h="center" if j > 0 else "left")
        ws.row_dimensions[row].height = 17

    for j, v in enumerate(["Итого", "", "", "", len(records)]):
        c = ws.cell(row=17, column=j+2, value=v)
        sc(c, bg=C_GRAY_HEAD, fg=C_WHITE, bold=True, size=10,
           h="center" if j > 0 else "left")
    ws.row_dimensions[17].height = 17

    ws["B19"].value = "Распределение по категориям"
    sc(ws["B19"], fg=C_DARK, bold=True, size=11, border=False)
    ws.row_dimensions[19].height = 22

    pie_data = [
        ("Встроены в репозиторий", counts["vendored / in-tree"]),
        ("Внешние зависимости",    counts["external system dependency"]),
        ("Не определено",          counts["unresolved"]),
    ]
    pie_start = 21
    for i, (label, value) in enumerate(pie_data):
        ws.cell(row=pie_start+i, column=2, value=label)
        ws.cell(row=pie_start+i, column=3, value=value)
        ws.row_dimensions[pie_start+i].height = 0

    pie = PieChart()
    pie.title = None; pie.style = 10; pie.width = 12; pie.height = 10
    labels = Reference(ws, min_col=2, min_row=pie_start, max_row=pie_start+2)
    data   = Reference(ws, min_col=3, min_row=pie_start, max_row=pie_start+2)
    pie.add_data(data); pie.set_categories(labels)
    for idx, color in enumerate([C_GREEN_HEAD, C_BLUE_HEAD, C_AMBER_HEAD]):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        pie.series[0].dPt.append(pt)
    ws.add_chart(pie, "F13")

    # ── Sheet 2: Component Inventory ─────────────────────────────────────────
    ws2 = wb.create_sheet("Инвентарь компонентов")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"

    inv_cols = [
        ("Компонент",              22), ("Классификация",        26),
        ("Версия",                 16), ("Уверенность в версии", 14),
        ("CPE",                    46), ("Уровень улик",         30),
        ("Уверенность",            12), ("Опционально / платформа",28),
        ("Обоснование",            54), ("Чего не хватает",      36),
        ("Действие",               38),
    ]
    for j, (name, width) in enumerate(inv_cols, 1):
        ws2.column_dimensions[col(j)].width = width

    ws2.merge_cells(f"A1:{col(len(inv_cols))}1")
    c = ws2["A1"]; c.value = "Инвентарь компонентов — все классифицированные зависимости"
    sc(c, bg=C_DARK, fg=C_WHITE, bold=True, size=12, h="left", border=False)
    ws2.row_dimensions[1].height = 28

    for j, (name, _) in enumerate(inv_cols, 1):
        c = ws2.cell(row=2, column=j, value=name)
        sc(c, bg=C_BLUE_HEAD, fg=C_WHITE, bold=True, size=10, h="center")
        if name in _COL_COMMENTS:
            _add_comment(c, _COL_COMMENTS[name])
    ws2.row_dimensions[2].height = 20

    for i, rec in enumerate(interesting):
        row = i + 3
        bg  = C_WHITE if i % 2 == 0 else "F8FAFC"
        action = _action_for(rec)
        ver_conf_color = {
            "high":    (C_GREEN_LIGHT,  C_GREEN_HEAD),
            "medium":  (C_AMBER_LIGHT,  C_AMBER_HEAD),
            "low":     (C_RED_LIGHT,    C_RED_HEAD),
        }.get(rec.version_confidence, (bg, "000000"))
        vals = [
            rec.normalized_name,
            rec.classification,
            rec.version if rec.version else "—",
            rec.version_confidence.upper() if rec.version_confidence else "—",
            rec.cpe if rec.cpe else "—",
            rec.evidence_level,
            rec.confidence.upper(),
            rec.optional_or_platform_specific if rec.optional_or_platform_specific != "no" else "—",
            rec.why,
            "; ".join(rec.missing_evidence) if rec.missing_evidence else "—",
            action,
        ]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(row=row, column=j, value=v)
            if j == 2:
                sc(c, bg=CLS_BG.get(rec.classification, bg),
                   fg=CLS_FG.get(rec.classification, "000000"), bold=True, size=10)
            elif j == 3:
                sc(c, bg=bg, bold=True, size=10,
                   fg=C_GREEN_HEAD if rec.version not in ("unknown", "", "n/a") else "888888")
            elif j == 4:
                sc(c, bg=ver_conf_color[0], fg=ver_conf_color[1],
                   bold=True, size=9, h="center")
            elif j == 5:
                sc(c, bg=bg, size=8, wrap=False,
                   fg="154360" if rec.cpe and "*" not in rec.cpe.split(":")[5] else "888888")
            elif j == 7:
                sc(c, bg=CONF_BG.get(rec.confidence, bg),
                   fg=CONF_FG.get(rec.confidence, "000000"), bold=True, size=10, h="center")
            elif j in (9, 10, 11):
                sc(c, bg=bg, size=9, wrap=True, fg="444444")
            else:
                sc(c, bg=bg, size=10, bold=(j==1))
        ws2.row_dimensions[row].height = max(20, min(60, 14 + (len(rec.why) // 60) * 14))

    tab2 = Table(displayName="ComponentTable",
                 ref=f"A2:{col(len(inv_cols))}{len(interesting)+2}")
    tab2.tableStyleInfo = TableStyleInfo(name="TableStyleLight2", showRowStripes=True)
    ws2.add_table(tab2)

    # ── Sheet 3: Evidence Details ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Улики")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A3"

    ev_cols_cfg = [
        ("Компонент", 18), ("Классификация", 26), ("Тип улики", 24),
        ("Файл", 40),      ("Строка", 8),          ("Фрагмент кода", 60),
    ]
    for j, (name, width) in enumerate(ev_cols_cfg, 1):
        ws3.column_dimensions[col(j)].width = width

    ws3.merge_cells(f"A1:{col(len(ev_cols_cfg))}1")
    c = ws3["A1"]; c.value = "Улики — все собранные доказательства по каждому компоненту"
    sc(c, bg=C_DARK, fg=C_WHITE, bold=True, size=12, h="left", border=False)
    ws3.row_dimensions[1].height = 28

    for j, (name, _) in enumerate(ev_cols_cfg, 1):
        c = ws3.cell(row=2, column=j, value=name)
        sc(c, bg=C_BLUE_HEAD, fg=C_WHITE, bold=True, size=10, h="center")
        if name in _COL_COMMENTS:
            _add_comment(c, _COL_COMMENTS[name])
    ws3.row_dimensions[2].height = 20

    ev_row = 3; global_i = 0
    for rec in interesting:
        for ev in rec.all_evidence():
            row_bg = "F8FAFC" if global_i % 2 == 0 else C_WHITE
            ev_bg  = EV_BG.get(ev.kind, row_bg)
            ev_fg  = EV_FG.get(ev.kind, "000000")
            vals = [
                rec.normalized_name, rec.classification,
                _ev_kind_ru(ev.kind),
                ev.file,
                ev.line if ev.line is not None else "—",
                ev.snippet[:300],
            ]
            for j, v in enumerate(vals, 1):
                c = ws3.cell(row=ev_row, column=j, value=v)
                if j == 2:
                    sc(c, bg=CLS_BG.get(rec.classification, row_bg),
                       fg=CLS_FG.get(rec.classification, "000000"), size=9)
                elif j == 3:
                    sc(c, bg=ev_bg, fg=ev_fg, bold=True, size=9, h="center")
                elif j == 6:
                    sc(c, bg=row_bg, size=9, wrap=True, fg="444444")
                else:
                    sc(c, bg=row_bg, size=9, bold=(j==1))
            ws3.row_dimensions[ev_row].height = 32 if len(ev.snippet) > 80 else 18
            ev_row += 1; global_i += 1

    if ev_row > 3:
        tab3 = Table(displayName="EvidenceTable",
                     ref=f"A2:{col(len(ev_cols_cfg))}{ev_row-1}")
        tab3.tableStyleInfo = TableStyleInfo(name="TableStyleLight2", showRowStripes=True)
        ws3.add_table(tab3)

    # ── Sheet 4: Optional Backends ────────────────────────────────────────────
    ws4 = wb.create_sheet("Опциональные бэкенды")
    ws4.sheet_view.showGridLines = False

    opt_cols_cfg = [
        ("Компонент", 20), ("Тип бэкенда", 36), ("Классификация", 26),
        ("Уверенность", 12), ("Обоснование", 54),
    ]
    for j, (name, width) in enumerate(opt_cols_cfg, 1):
        ws4.column_dimensions[col(j)].width = width

    ws4.merge_cells(f"A1:{col(len(opt_cols_cfg))}1")
    c = ws4["A1"]; c.value = "Опциональные и платформозависимые бэкенды"
    sc(c, bg=C_DARK, fg=C_WHITE, bold=True, size=12, h="left", border=False)
    ws4.row_dimensions[1].height = 28

    for j, (name, _) in enumerate(opt_cols_cfg, 1):
        c = ws4.cell(row=2, column=j, value=name)
        sc(c, bg="6C3483", fg=C_WHITE, bold=True, size=10, h="center")
        if name in _COL_COMMENTS:
            _add_comment(c, _COL_COMMENTS[name])
    ws4.row_dimensions[2].height = 20

    if optional_rec:
        for i, rec in enumerate(optional_rec):
            row = i + 3
            bg  = "FAF0FF" if i % 2 == 0 else C_WHITE
            vals = [
                rec.normalized_name,
                rec.optional_or_platform_specific.replace("yes: ", ""),
                rec.classification,
                rec.confidence.upper(),
                rec.why,
            ]
            for j, v in enumerate(vals, 1):
                c = ws4.cell(row=row, column=j, value=v)
                if j == 3:
                    sc(c, bg=CLS_BG.get(rec.classification, bg),
                       fg=CLS_FG.get(rec.classification, "000000"), size=10)
                elif j == 4:
                    sc(c, bg=CONF_BG.get(rec.confidence, bg),
                       fg=CONF_FG.get(rec.confidence, "000000"), bold=True, size=10, h="center")
                elif j == 5:
                    sc(c, bg=bg, size=9, wrap=True, fg="444444")
                else:
                    sc(c, bg=bg, size=10, bold=(j==1))
            ws4.row_dimensions[row].height = 36
    else:
        c = ws4.cell(row=3, column=1, value="Опциональные бэкенды не обнаружены.")
        sc(c, fg="888888", size=10, italic=True)

    # ── Sheet 5: System Headers ───────────────────────────────────────────────
    ws5 = wb.create_sheet("Системные заголовки")
    ws5.sheet_view.showGridLines = False

    sys_cols_cfg = [("Заголовок / интринсик", 32), ("Категория", 28), ("Причина", 44)]
    for j, (name, width) in enumerate(sys_cols_cfg, 1):
        ws5.column_dimensions[col(j)].width = width

    ws5.merge_cells(f"A1:{col(len(sys_cols_cfg))}1")
    c = ws5["A1"]
    c.value = f"Системные и компиляторные заголовки — {len(system_recs)} шт. (не сторонние зависимости)"
    sc(c, bg=C_DARK, fg=C_WHITE, bold=True, size=12, h="left", border=False)
    ws5.row_dimensions[1].height = 28

    for j, (name, _) in enumerate(sys_cols_cfg, 1):
        c = ws5.cell(row=2, column=j, value=name)
        sc(c, bg=C_GRAY_HEAD, fg=C_WHITE, bold=True, size=10, h="center")
        if name in _COL_COMMENTS:
            _add_comment(c, _COL_COMMENTS[name])
    ws5.row_dimensions[2].height = 20

    for i, rec in enumerate(system_recs):
        row = i + 3
        bg  = C_WHITE if i % 2 == 0 else C_GRAY_LIGHT
        if rec.is_compiler_intrinsic:
            category = "интринсик компилятора"
        elif rec.is_windows_specific:
            category = "Windows SDK"
        elif rec.is_kernel_or_sdk_header:
            category = "заголовок ядра / платформы"
        else:
            category = "стандартная библиотека"
        vals = [
            rec.normalized_name, category,
            rec.why or "Стандартный/системный заголовок, не сторонняя зависимость.",
        ]
        for j, v in enumerate(vals, 1):
            c = ws5.cell(row=row, column=j, value=v)
            sc(c, bg=bg, size=9, fg="555555" if j > 1 else "000000", bold=(j==1))
        ws5.row_dimensions[row].height = 16

    # ── Sheets 6-9: Segmented by discovery_source ────────────────────────────
    segments = [
        ("in-tree source",          "Исходники в репо",   C_GREEN_HEAD, C_GREEN_LIGHT,
         "Компоненты, найденные как встроенные директории с исходным кодом (.c/.cpp) внутри репозитория."),
        ("build config",            "Конфиг сборки",      C_BLUE_HEAD,  C_BLUE_LIGHT,
         "Компоненты, найденные только в файлах сборки (configure.ac, Makefile.am, CMakeLists.txt и т.д.)."),
        ("header only",             "Только заголовки",   C_AMBER_HEAD, C_AMBER_LIGHT,
         "Компоненты найдены только через #include — улик из системы сборки нет. Требуют проверки."),
        ("in-tree source + build config", "Исходники + конфиг", C_GREEN_HEAD, "EAF9F0",
         "Компоненты с исходниками внутри репозитория И подтверждением в системе сборки — максимальная уверенность."),
    ]

    seg_cols_cfg = [
        ("Компонент",              20),
        ("Классификация",          26),
        ("Уровень улик",           28),
        ("Уверенность",            12),
        ("Источник обнаружения",   24),
        ("Опционально / платформа",24),
        ("Обоснование",            50),
        ("Действие",               38),
    ]

    for seg_key, seg_title, hc, lc, seg_desc in segments:
        if seg_key in ("in-tree source + build config",):
            subset = [r for r in records if r.discovery_source == seg_key]
        else:
            subset = [r for r in records
                      if r.discovery_source == seg_key
                      or (seg_key != "in-tree source + build config"
                          and r.discovery_source.startswith(seg_key)
                          and "+" not in r.discovery_source)]

        safe_title = seg_title[:31]
        wsN = wb.create_sheet(safe_title)
        wsN.sheet_view.showGridLines = False
        wsN.freeze_panes = "A3"

        for j, (name, width) in enumerate(seg_cols_cfg, 1):
            wsN.column_dimensions[col(j)].width = width

        wsN.merge_cells(f"A1:{col(len(seg_cols_cfg))}1")
        c = wsN["A1"]
        c.value = f"{seg_title} ({len(subset)} компонентов) — {seg_desc}"
        sc(c, bg=C_DARK, fg=C_WHITE, bold=True, size=11, h="left", border=False)
        wsN.row_dimensions[1].height = 28

        for j, (name, _) in enumerate(seg_cols_cfg, 1):
            c = wsN.cell(row=2, column=j, value=name)
            sc(c, bg=hc, fg=C_WHITE, bold=True, size=10, h="center")
            if name in _COL_COMMENTS:
                _add_comment(c, _COL_COMMENTS[name])
        wsN.row_dimensions[2].height = 20

        if not subset:
            c = wsN.cell(row=3, column=1, value="В этом сегменте компонентов нет.")
            sc(c, fg="888888", size=10, italic=True)
            continue

        for i, rec in enumerate(subset):
            row = i + 3
            bg  = lc if i % 2 == 0 else C_WHITE
            action = _action_for(rec)
            vals = [
                rec.normalized_name,
                rec.classification,
                rec.evidence_level,
                rec.confidence.upper(),
                rec.discovery_source,
                rec.optional_or_platform_specific if rec.optional_or_platform_specific != "no" else "-",
                rec.why,
                action,
            ]
            for j, v in enumerate(vals, 1):
                c = wsN.cell(row=row, column=j, value=v)
                if j == 2:
                    sc(c, bg=CLS_BG.get(rec.classification, bg),
                       fg=CLS_FG.get(rec.classification, "000000"), bold=True, size=10)
                elif j == 4:
                    sc(c, bg=CONF_BG.get(rec.confidence, bg),
                       fg=CONF_FG.get(rec.confidence, "000000"), bold=True, size=10, h="center")
                elif j in (7, 8):
                    sc(c, bg=bg, size=9, wrap=True, fg="444444")
                else:
                    sc(c, bg=bg, size=10, bold=(j==1))
            wsN.row_dimensions[row].height = max(20, min(60, 14 + (len(rec.why) // 60) * 14))

        seg_table = Table(
            displayName=f"Seg_{seg_title.replace(' ', '_').replace('+', 'And')}",
            ref=f"A2:{col(len(seg_cols_cfg))}{len(subset)+2}",
        )
        seg_table.tableStyleInfo = TableStyleInfo(name="TableStyleLight2", showRowStripes=True)
        wsN.add_table(seg_table)

    # ── Sheet 10: Glossary ────────────────────────────────────────────────────
    wsg = wb.create_sheet("Глоссарий")
    wsg.sheet_view.showGridLines = False

    glossary_cols = [("Термин", 32), ("Значение", 70)]
    for j, (name, width) in enumerate(glossary_cols, 1):
        wsg.column_dimensions[col(j)].width = width

    wsg.merge_cells("A1:B1")
    c = wsg["A1"]; c.value = "Глоссарий — расшифровка всех терминов отчёта"
    sc(c, bg=C_DARK, fg=C_WHITE, bold=True, size=12, h="left", border=False)
    wsg.row_dimensions[1].height = 28

    for j, (name, _) in enumerate(glossary_cols, 1):
        c = wsg.cell(row=2, column=j, value=name)
        sc(c, bg=C_GRAY_HEAD, fg=C_WHITE, bold=True, size=10, h="center")
    wsg.row_dimensions[2].height = 20

    glossary = [
        # Классификации
        ("vendored / in-tree",
         "Компонент встроен в репозиторий: его исходный код (.c/.cpp/.h) лежит прямо внутри проекта и собирается вместе с ним. Не нужно устанавливать на хост."),
        ("external system dependency",
         "Внешняя зависимость: библиотека должна быть установлена на хост-машине перед сборкой. Агент нашёл её упоминание в системе сборки или при линковке."),
        ("unresolved",
         "Не удалось классифицировать: данных недостаточно. Обычно найден только #include без подтверждения из системы сборки. Требует ручной проверки."),
        ("not a library",
         "Не является сторонней библиотекой: это стандартный заголовок C/C++, системный заголовок POSIX, заголовок ядра или интринсик компилятора."),
        # Уровни улик
        ("in-tree source + build participation",
         "Сильнейший уровень: найдены исходники внутри репо И подтверждение в системе сборки (LDADD, target_link_libraries и т.д.)."),
        ("confirmed linked",
         "Подтверждена линковка: компонент явно указан в шаге финальной линковки (-lNAME или target_link_libraries). Высокая уверенность."),
        ("build-integrated",
         "Встроен в сборку: компонент передаётся в переменные сборки (AC_SUBST, AM_CONDITIONAL, LDADD). Средняя уверенность."),
        ("probe only",
         "Только проверка наличия: система сборки ищет библиотеку (AC_CHECK_LIB, PKG_CHECK_MODULES), но интеграция в сборку не подтверждена. Низкая уверенность."),
        ("insufficient evidence",
         "Недостаточно улик: найден только #include или ничего. Классификация невозможна без дополнительных данных."),
        ("system/platform/compiler header",
         "Системный/платформенный/компиляторный заголовок — не является зависимостью третьей стороны."),
        # Уверенность
        ("HIGH (высокая)",
         "Агент уверен в результате. Несколько независимых улик согласуются между собой. Ручная проверка не требуется."),
        ("MEDIUM (средняя)",
         "Агент достаточно уверен, но одна из улик отсутствует или неоднозначна. Рекомендуется беглая проверка."),
        ("LOW (низкая)",
         "Низкая уверенность. Улика только одного слабого типа (#include или probe). Обязательно проверить вручную."),
        # Типы улик
        ("Улика: исходники в репо (in tree source)",
         "Найдена директория с файлами .c/.cpp/.h внутри репозитория. Сильный признак vendored-зависимости."),
        ("Улика: финальная линковка (final link)",
         "Компонент упомянут в target_link_libraries, LDADD, LIBADD или через флаг -lNAME. Самая сильная build-улика."),
        ("Улика: интеграция в сборку (build integration)",
         "Компонент передаётся в переменные сборки через AC_SUBST, AM_CONDITIONAL и т.д. Средняя сила."),
        ("Улика: обнаружение (build detection)",
         "Система сборки проверяет наличие компонента через AC_CHECK_LIB, PKG_CHECK_MODULES, find_package. Слабая улика — только проверка, не использование."),
        ("Улика: include",
         "#include <libname.h> найден в исходном коде. Самая слабая улика — не доказывает что библиотека реально используется или слинкована."),
        ("Улика: платформенная заметка (platform note)",
         "Упоминание платформы (windows, linux, dpdk, pfring) рядом с именем компонента. Справочная информация для колонки 'Опционально'."),
        # Источник обнаружения
        ("Источник: in-tree source",
         "Компонент обнаружен через директорию с исходниками внутри репозитория."),
        ("Источник: build config",
         "Компонент обнаружен только через файлы системы сборки (configure.ac, CMakeLists.txt, Makefile.am и т.д.)."),
        ("Источник: header only",
         "Компонент обнаружен только через директивы #include. Улик из системы сборки нет."),
        ("Источник: in-tree source + build config",
         "Компонент обнаружен и через исходники внутри репо, и через файлы сборки. Максимальная уверенность."),
        ("Источник: system header",
         "Системный или компиляторный заголовок. Не является сторонней зависимостью."),
        # Действия
        ("Действие: нет действий",
         "Компонент встроен в репозиторий, уверенность высокая. Дополнительных шагов не требуется."),
        ("Действие: установить на хост",
         "Внешняя зависимость с высокой уверенностью. Необходимо установить перед сборкой (apt install, brew install и т.д.)."),
        ("Действие: проверить вручную",
         "Низкая уверенность или недостаточно улик. Рекомендуется открыть лист 'Улики' и проверить файлы вручную."),
        ("Действие: уточнить источник",
         "Исходники найдены, но участие в сборке не подтверждено. Проверить Makefile.am / CMakeLists.txt."),
    ]

    for i, (term, meaning) in enumerate(glossary):
        row = i + 3
        bg  = C_WHITE if i % 2 == 0 else C_GRAY_LIGHT

        # section separator rows
        is_section = term.startswith("Улика:") or term.startswith("Источник:") or term.startswith("Действие:")
        sep_before = (
            term == "in-tree source + build participation" or
            term == "HIGH (высокая)" or
            term == "Улика: исходники в репо (in tree source)" or
            term == "Источник: in-tree source" or
            term == "Действие: нет действий"
        )
        if sep_before:
            # blank separator row
            for jj in range(1, 3):
                wsg.cell(row=row, column=jj).fill = _fill("E8EEF4")
            wsg.row_dimensions[row].height = 6
            row += 1
            wsg.row_dimensions[row].height = 18
            bg = "EEF4FF"

        c1 = wsg.cell(row=row, column=1, value=term)
        sc(c1, bg=bg, size=10, bold=True, wrap=False)
        c2 = wsg.cell(row=row, column=2, value=meaning)
        sc(c2, bg=bg, size=9, wrap=True, fg="444444")
        wsg.row_dimensions[row].height = max(18, min(50, 14 + (len(meaning) // 80) * 14))

    wb.save(str(out_path))


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def write_reports(records: list[ComponentRecord], out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    json_path = out_dir / "report.json"
    md_path   = out_dir / "report.md"
    xlsx_path = out_dir / "report.xlsx"

    payload = {
        "summary": {cls: sum(r.classification == cls for r in records) for cls in _CLASSIFICATIONS},
        "components": [r.to_dict() for r in records],
    }

    json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    md_path.write_text(build_markdown(records), encoding="utf-8")
    print(f"  report.json  -> {json_path}")
    print(f"  report.md    -> {md_path}")

    build_xlsx(records, xlsx_path)
    print(f"  report.xlsx  -> {xlsx_path}")

    # CycloneDX SBOM
    from sbom_writer import write_sbom
    repo_name = out_dir.parent.name if out_dir.parent else "unknown"
    path_known, path_unknown = write_sbom(records, out_dir, repo_name=repo_name)
    print(f"  sbom_known.json   -> {path_known}")
    print(f"  sbom_unknown.json -> {path_unknown}")
